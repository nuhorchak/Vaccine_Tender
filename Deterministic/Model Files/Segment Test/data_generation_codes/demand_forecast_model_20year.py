"""
Vaccine Demand Forecast Model
Replicates Demand_Scenarios_Base.xlsx logic

Structure:
- total_capacity_calcs: Historical + projected supply capacity by vaccine, income group (2000-2030)
- Medium_Demand_calc:   Antigen-level demand aggregations (vaccine format → antigen)
- Medium_Demand:        Final demand by antigen (FORECAST_PERIODS forecast periods)

Methodology:
- Historical data (2000-2022): Actual supply figures by income group
- Projections (2023-2030): Trend-based continuation using weighted averages
- Antigen aggregation: Vaccines mapped to antigens (e.g., MMR contributes to Measles, Mumps, Rubella)
- Market share splits: Producer-level allocations from vaccine-producer list
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# FORECAST CONFIGURATION — change only this value to adjust forecast length
# ─────────────────────────────────────────────────────────────────────────────

FORECAST_PERIODS = 20          # ← set number of forecast years here
FORECAST_START_YEAR = 2021
FORECAST_END_YEAR   = FORECAST_START_YEAR + FORECAST_PERIODS - 1
FORECAST_YEARS      = list(range(FORECAST_START_YEAR, FORECAST_START_YEAR + FORECAST_PERIODS))

# Derived Excel layout values (col A = labels, cols B onward = periods)
_LAST_PERIOD_COL    = FORECAST_PERIODS + 1          # e.g. 21 for 20 periods
_LAST_COL_LETTER    = get_column_letter(_LAST_PERIOD_COL)   # e.g. 'U'
_MERGE_RANGE        = f'A1:{_LAST_COL_LETTER}1'     # title merge range
_MERGE_RANGE_ROW2   = f'A2:{_LAST_COL_LETTER}2'    # subtitle merge range

# ─────────────────────────────────────────────────────────────────────────────
# 1. RAW DATA EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

SOURCE = 'Demand_Scenarios_Base.xlsx'

def load_source():
    sheets = pd.read_excel(SOURCE, sheet_name=None, header=None)
    return sheets

# ─────────────────────────────────────────────────────────────────────────────
# 2. VACCINE → ANTIGEN MAPPING
# ─────────────────────────────────────────────────────────────────────────────

ANTIGEN_MAP = {
    'Measles':         ['Measles'],
    'MR':              ['Measles', 'Rubella'],
    'MMR':             ['Measles', 'Mumps', 'Rubella'],
    'TT':              ['Tetanus'],
    'Td':              ['Tetanus', 'Diphtheria'],
    'DT':              ['Diphtheria', 'Tetanus'],
    'DTwP':            ['Diphtheria', 'Tetanus', 'Pertussis'],
    'DTaP':            ['Diphtheria', 'Tetanus', 'Pertussis'],
    'DTwP-Hib':        ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib'],
    'DTaP-Hib':        ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib'],
    'Penta':           ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B'],
    'DTwP-HepB-Hib':   ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B'],
    'Hexa':            ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B', 'IPV'],
    'DTaP-HepB-Hib-IPV': ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B', 'IPV'],
    'IPV':             ['IPV'],
    'OPV':             ['Polio'],
    'bOPV':            ['Polio'],
    'tOPV':            ['Polio'],
    'HPV':             ['HPV'],
    'HepB':            ['Hepatitis_B'],
    'Hib':             ['Hib'],
    'Rotavirus':       ['Rotavirus'],
    'Rota':            ['Rotavirus'],
    'PCV':             ['PCV'],
}

# Vaccine short codes as used in total_capacity_calcs
VACCINE_CODES = {
    'M':          'Measles',
    'MR':         'MR',
    'MMR':        'MMR',
    'TT':         'TT',
    'Td':         'Td',
    'DT':         'DT',
    'DTwP':       'DTwP',
    'DTwP-Hib':   'DTwP-Hib',
    'Penta':      'DTwP-HepB-Hib',
    'Hexa':       'DTaP-HepB-Hib-IPV',
    'IPV':        'IPV',
    'OPV':        'bOPV',
    'HPV':        'HPV',
    'HepB':       'HepB',
    'Hib':        'Hib',
    'Rotavirus':  'Rota',
    'PCV':        'PCV',
}

# ─────────────────────────────────────────────────────────────────────────────
# 3. MARKET SHARE TABLE
# ─────────────────────────────────────────────────────────────────────────────

MARKET_SHARES = {
    'M': {
        'Serum Institute of India Pvt. Ltd.': 0.8182,
        'PT Bio Farma (Persero)':              0.1818,
    },
    'MR': {
        'Serum Institute of India Pvt. Ltd.': 0.7241,
        'Biological E. Limited':               0.2759,
    },
    'MMR': {
        'GlaxoSmithKline Biologicals SA':    0.2921,
        'Serum Institute of India':           0.7079,
    },
    'TT': {
        'Serum Institute of India Pvt. Ltd.': 0.7412,
        'PT Bio Farma (Persero)':              0.1647,
        'BB- NCIPD Ltd. (Bulgaria)':           0.0941,
    },
    'Td': {
        'Serum Institute of India Pvt. Ltd.': 0.7412,
        'PT Bio Farma (Persero)':              0.1647,
        'BB- NCIPD Ltd. (Bulgaria)':           0.0941,
    },
    'DT': {
        'Serum Institute of India Pvt. Ltd.': 0.7412,
        'PT Bio Farma (Persero)':              0.1647,
        'BB- NCIPD Ltd. (Bulgaria)':           0.0941,
    },
    'DTwP': {
        'Serum Institute of India Pvt. Ltd.': 0.7241,
        'Biological E. Limited':               0.2759,
    },
    'Penta': {
        'Serum Institute of India Pvt. Ltd.': 0.5575,
        'PT Bio Farma (Persero)':              0.1239,
        'Biological E. Limited':               0.2124,
        'LG Chem Ltd':                         0.0708,
        'Panacea Biotec Ltd.':                 0.0354,
    },
    'HepB': {
        'Serum Institute of India Pvt. Ltd.': 0.8873,
        'LG Chem Ltd':                         0.1127,
    },
    'Hib': {
        'Serum Institute of India Pvt. Ltd.': 0.6923,
        'Sanofi Pasteur':                      0.2747,
        'Centro de Ingenieria Genetica y Biotecnologia': 0.0330,
    },
    'Rotavirus': {
        'Serum Institute of India Pvt. Ltd.': 0.6364,
        'Bharat Biotech International Limited': 0.0731,
        'GlaxoSmithKline Biologicals SA':       0.2626,
    },
    'PCV': {
        'Serum Institute of India Pvt. Ltd.': 0.6848,
        'Pfizer':                               0.0326,
        'GlaxoSmithKline Biologicals SA':       0.2826,
    },
    'HPV': {
        'GlaxoSmithKline Biologicals SA':    0.7879,
        'Merck Sharp & Dohme LLC':            0.1818,
        'Xiamen Innovax Biotech Co. Ltd.':    0.0303,
    },
    'IPV': {
        'LG Chem Ltd':                              0.3636,
        'AJ Vaccines A/S':                          0.1364,
        'Bilthoven Biologicals B.V.':               0.0909,
        'Sanofi Healthcare India Private Limited':  0.4091,
    },
    'OPV': {
        'Serum Institute of India Pvt. Ltd.': 0.4257,
        'PT Bio Farma (Persero)':              0.0946,
        'GlaxoSmithKline Biologicals SA':      0.1757,
        'Sanofi Pasteur':                      0.1689,
        'Panacea Biotec Ltd.':                 0.0270,
        'Beijing Institute of Biological Products Co., Ltd.': 0.0203,
        'Bharat Biotech International Limited': 0.0676,
        'Haffkine Bio Pharmaceutical Corporation Ltd': 0.0203,
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# 4. PARSE CAPACITY DATA
# ─────────────────────────────────────────────────────────────────────────────

def parse_capacity(sheets):
    """
    Parse total_capacity_calcs sheet.
    Returns DataFrame indexed by (Vaccine, wb_status), columns = years 2000–2030.
    """
    df = sheets['total_capacity_calcs']
    years = list(range(2000, 2031))

    # Header row 0: col 0=blank, col 1=Vaccine, col 2=wb_status, cols 3–33=years
    records = []
    for _, row in df.iterrows():
        vaccine = row.iloc[1]
        status  = row.iloc[2]
        if pd.isna(vaccine) or vaccine == 'Vaccine':
            continue
        vals = {}
        for i, yr in enumerate(years):
            v = row.iloc[3 + i]
            vals[yr] = float(v) if pd.notna(v) else np.nan
        records.append({'vaccine': str(vaccine).strip(),
                        'status':  str(status).strip(),
                        **vals})

    cap = pd.DataFrame(records).set_index(['vaccine', 'status'])
    return cap

# ─────────────────────────────────────────────────────────────────────────────
# 5. PARSE MEDIUM DEMAND DATA (final output)
# ─────────────────────────────────────────────────────────────────────────────

def parse_medium_demand(sheets):
    """Parse the Medium_Demand sheet — reads exactly FORECAST_PERIODS columns."""
    df = sheets['Medium_Demand']
    data = {}
    for _, row in df.iterrows():
        label = row.iloc[0]
        if pd.isna(label):
            continue
        vals = [row.iloc[i] for i in range(1, FORECAST_PERIODS + 1)]
        if pd.notna(label) and isinstance(label, str) and not label.startswith('1'):
            data[label] = [float(v) if pd.notna(v) else 0.0 for v in vals]
    return pd.DataFrame(data, index=range(1, FORECAST_PERIODS + 1)).T

# ─────────────────────────────────────────────────────────────────────────────
# 6. BUILD FORECAST: rolling average of last N years by income group
# ─────────────────────────────────────────────────────────────────────────────

INCOME_GROUPS = ['High Income', 'Upper Middle Income', 'Lower Middle Income', 'Low Income', '--']

HIST_WINDOW = 5  # years to use for trend/average

def project_series(historical_vals, n_periods=FORECAST_PERIODS, window=HIST_WINDOW):
    """
    Project forward n_periods using a weighted moving average of the last `window` years.
    Weights: most recent year gets highest weight.
    """
    vals = [v for v in historical_vals if not np.isnan(v)]
    if len(vals) == 0:
        return [0.0] * n_periods
    recent = vals[-window:] if len(vals) >= window else vals
    weights = np.arange(1, len(recent) + 1, dtype=float)
    weights /= weights.sum()
    avg = float(np.dot(weights, recent))

    # Compute trend (linear slope over recent window)
    if len(recent) >= 3:
        x = np.arange(len(recent), dtype=float)
        slope = np.polyfit(x, recent, 1)[0]
    else:
        slope = 0.0

    projected = []
    for i in range(n_periods):
        val = avg + slope * (i + 1) * 0.3   # dampened trend
        val = max(val, 0.0)
        projected.append(val)
    return projected

def build_capacity_projections(cap_df):
    """
    For each (vaccine, status) pair, project values for forecast years
    using historical data through 2020 (available) + trend extrapolation.
    """
    hist_years = list(range(2000, 2021))
    projections = {}

    for (vaccine, status), row in cap_df.iterrows():
        hist = [row.get(y, np.nan) for y in hist_years]
        proj = project_series(hist, n_periods=FORECAST_PERIODS, window=HIST_WINDOW)
        projections[(vaccine, status)] = proj

    return projections

# ─────────────────────────────────────────────────────────────────────────────
# 7. AGGREGATE VACCINE → ANTIGEN DEMAND
# ─────────────────────────────────────────────────────────────────────────────

DEMAND_VACCINE_MAP = {
    'M':          ['Measles'],
    'MR':         ['MR'],
    'MMR':        ['MMR', 'MMRV'],
    'TT':         ['TT'],
    'Td':         ['Td'],
    'DT':         ['DT'],
    'DTwP':       ['DTwP'],
    'DTwP-Hib':   ['DTwP-Hib'],
    'Penta':      ['DTwP-HepB-Hib'],
    'Hexa':       ['DTaP-HepB-Hib-IPV'],
    'IPV':        ['IPV', 'DTaP-HepB-Hib-IPV', 'DTaP-Hib-IPV', 'DTaP-HepB-IPV'],
    'OPV':        ['bOPV', 'tOPV'],
    'HPV':        ['HPV'],
    'HepB':       ['HepB'],
    'Hib':        ['Hib'],
    'Rotavirus':  ['Rota'],
    'PCV':        ['PCV'],
}

ANTIGEN_FROM_VACCINE_CODE = {
    'Measles':   ['Measles'],
    'MR':        ['Measles', 'Rubella'],
    'MMR':       ['Measles', 'Mumps', 'Rubella'],
    'TT':        ['Tetanus'],
    'Td':        ['Diphtheria', 'Tetanus'],
    'DT':        ['Diphtheria', 'Tetanus'],
    'DTwP':      ['Diphtheria', 'Tetanus', 'Pertussis'],
    'DTwP-Hib':  ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib'],
    'Penta':     ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B'],
    'Hexa':      ['Diphtheria', 'Tetanus', 'Pertussis', 'Hib', 'Hepatitis_B', 'IPV'],
    'IPV':       ['IPV'],
    'OPV':       ['Polio'],
    'HPV':       ['HPV'],
    'HepB':      ['Hepatitis_B'],
    'Hib':       ['Hib'],
    'Rotavirus': ['Rotavirus'],
    'PCV':       ['PCV'],
}

ALL_ANTIGENS = ['Measles', 'Mumps', 'Rubella', 'Diphtheria', 'Tetanus',
                'Pertussis', 'Hepatitis_B', 'Hib', 'IPV', 'Polio',
                'HPV', 'Rotavirus', 'PCV']

def aggregate_demand_by_antigen(cap_df, projections):
    """
    Sum projected demand across all contributing vaccines for each antigen.
    Returns dict: antigen -> list of FORECAST_PERIODS projected values.
    """
    antigen_demand = {a: np.zeros(FORECAST_PERIODS) for a in ALL_ANTIGENS}

    all_vaccines = cap_df.index.get_level_values('vaccine').unique()

    for vcode, contributing_vaccines in DEMAND_VACCINE_MAP.items():
        antigens = ANTIGEN_FROM_VACCINE_CODE.get(vcode, [])

        total_proj = np.zeros(FORECAST_PERIODS)
        for vaccine_name in contributing_vaccines:
            for income_group in INCOME_GROUPS:
                key = (vaccine_name, income_group)
                if key in projections:
                    proj = np.array(projections[key])
                    total_proj += proj

        for antigen in antigens:
            if antigen in antigen_demand:
                antigen_demand[antigen] += total_proj

    return antigen_demand

# ─────────────────────────────────────────────────────────────────────────────
# 8. SCENARIO SCALING FACTORS
# ─────────────────────────────────────────────────────────────────────────────

SCENARIOS = {
    'Low':    0.85,
    'Medium': 1.00,
    'High':   1.15,
}

def apply_scenarios(base_demand):
    """Apply scenario multipliers to base (Medium) demand."""
    results = {}
    for scenario, factor in SCENARIOS.items():
        results[scenario] = {
            antigen: [v * factor for v in vals]
            for antigen, vals in base_demand.items()
        }
    return results

# ─────────────────────────────────────────────────────────────────────────────
# 9. BUILD EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

# Color palette
BLUE_FONT    = 'FF0000FF'
BLACK_FONT   = 'FF000000'
GREEN_FONT   = 'FF008000'
HEADER_FILL  = 'FF1F4E79'
ALT_FILL     = 'FFD6E4F0'
WHITE_FILL   = 'FFFFFFFF'
SECTION_FILL = 'FFD9E1F2'

def make_font(bold=False, color=BLACK_FONT, size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def make_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def header_style(ws, cell_ref, value, bold=True):
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(bold=bold, color='FFFFFFFF', size=10, name='Arial')
    cell.fill = make_fill(HEADER_FILL)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def num_fmt(cell, fmt='#,##0'):
    cell.number_format = fmt

def border_thin():
    thin = Side(style='thin', color='FFB8CCE4')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_table_style(ws, row, col, value, is_alt=False, blue_input=False, formula=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = border_thin()
    cell.alignment = Alignment(horizontal='right', vertical='center')
    if blue_input:
        cell.font = make_font(color=BLUE_FONT)
    elif formula:
        cell.font = make_font(color=BLACK_FONT)
    else:
        cell.font = make_font()
    if is_alt:
        cell.fill = make_fill(ALT_FILL)
    else:
        cell.fill = make_fill(WHITE_FILL)
    cell.number_format = '#,##0'
    return cell

def build_excel(sheets, cap_df, projections, antigen_demand, scenario_data, actual_md):
    wb = Workbook()
    wb.remove(wb.active)

    _copy_reference_sheet(wb, sheets, 'vaccine-producer list', 'Vaccine-Producer List')
    _copy_reference_sheet(wb, sheets, 'vaccine-antigen list',  'Vaccine-Antigen List')
    _build_capacity_sheet(wb, cap_df, projections)
    _build_demand_calc_sheet(wb, antigen_demand, actual_md)
    _build_medium_demand_sheet(wb, scenario_data, actual_md)

    return wb


def _copy_reference_sheet(wb, sheets, src_name, dst_name):
    ws = wb.create_sheet(dst_name)
    df = sheets[src_name]
    ws.freeze_panes = 'B2'

    for r_idx, row in enumerate(df.values, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            if r_idx == 1:
                cell.font = Font(bold=True, color='FFFFFFFF', name='Arial', size=9)
                cell.fill = make_fill(HEADER_FILL)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            elif c_idx == 1:
                cell.font = Font(bold=True, name='Arial', size=9)
                cell.fill = make_fill(SECTION_FILL)
            else:
                cell.font = Font(name='Arial', size=9)
                if r_idx % 2 == 0:
                    cell.fill = make_fill(ALT_FILL)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)


def _build_capacity_sheet(wb, cap_df, projections):
    ws = wb.create_sheet('Capacity_Calcs')
    ws.freeze_panes = 'D2'

    years = list(range(2000, 2031))
    headers = ['Vaccine', 'Income Group'] + [str(y) for y in years]

    ws.merge_cells('A1:AJ1')
    title = ws['A1']
    title.value = 'Total Capacity Calculations by Vaccine & Income Group (2000–2030)'
    title.font = Font(bold=True, color='FFFFFFFF', size=12, name='Arial')
    title.fill = make_fill(HEADER_FILL)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF', size=9, name='Arial')
        cell.fill = make_fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border_thin()

    cur_vaccine = None
    for r_idx, ((vaccine, status), row) in enumerate(cap_df.iterrows(), start=3):
        is_alt = r_idx % 2 == 0
        bg = ALT_FILL if is_alt else WHITE_FILL

        if vaccine != cur_vaccine:
            cur_vaccine = vaccine
            ws.cell(row=r_idx, column=1, value=vaccine).font = Font(bold=True, size=9, name='Arial')
        else:
            ws.cell(row=r_idx, column=1, value='').fill = make_fill(bg)

        ws.cell(row=r_idx, column=2, value=status)

        for c_idx, yr in enumerate(years, start=3):
            val = row.get(yr, None)
            cell = ws.cell(row=r_idx, column=c_idx,
                           value=val if (pd.notna(val) if val is not None else False) else None)
            cell.number_format = '#,##0'
            cell.font = Font(size=9, name='Arial',
                             color=BLUE_FONT if c_idx <= 25 else BLACK_FONT)
            if is_alt:
                cell.fill = make_fill(ALT_FILL)
            cell.border = border_thin()
            cell.alignment = Alignment(horizontal='right')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1,
            value='Note: Blue = historical input; Black = formula projection').font = \
        Font(italic=True, size=8, color='FF595959', name='Arial')


def _build_demand_calc_sheet(wb, antigen_demand, actual_md):
    ws = wb.create_sheet('Medium_Demand_Calc')
    ws.freeze_panes = 'B3'

    periods = list(range(1, FORECAST_PERIODS + 1))

    ws.merge_cells(_MERGE_RANGE)
    t = ws['A1']
    t.value = 'Medium Demand Calculation – Vaccine Format Aggregations'
    t.font = Font(bold=True, color='FFFFFFFF', size=12, name='Arial')
    t.fill = make_fill(HEADER_FILL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value='Vaccine / Antigen').font = \
        Font(bold=True, color='FFFFFFFF', size=9, name='Arial')
    ws.cell(row=2, column=1).fill = make_fill(HEADER_FILL)
    for p in periods:
        cell = ws.cell(row=2, column=p + 1, value=f'Period {p}')
        cell.font = Font(bold=True, color='FFFFFFFF', size=9, name='Arial')
        cell.fill = make_fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal='center')

    vac_order = ['M', 'MR', 'MMR', 'TT', 'Td', 'DT', 'DTwP', 'DTwP-Hib',
                 'Penta', 'Hexa', 'IPV', 'OPV', 'HPV', 'HepB', 'Hib', 'Rotavirus', 'PCV']

    src = pd.read_excel(SOURCE, sheet_name='Medium_Demand_calc', header=None)
    actual_vac = {}
    for _, row in src.iterrows():
        label = row.iloc[0]
        if pd.isna(label) or str(label).strip() in ('', '1'):
            continue
        label = str(label).strip()
        vals = [float(row.iloc[i]) if pd.notna(row.iloc[i]) else 0.0
                for i in range(1, FORECAST_PERIODS + 1)]
        actual_vac[label] = vals

    row_num = 3
    ws.cell(row=row_num, column=1, value='VACCINE FORMAT DEMAND').font = \
        Font(bold=True, size=9, color=BLUE_FONT, name='Arial')
    ws.cell(row=row_num, column=1).fill = make_fill(SECTION_FILL)
    row_num += 1

    for vac in vac_order:
        is_alt = row_num % 2 == 0
        cell = ws.cell(row=row_num, column=1, value=vac)
        cell.font = Font(bold=True, size=9, name='Arial')
        cell.fill = make_fill(SECTION_FILL if not is_alt else ALT_FILL)

        vals = actual_vac.get(vac, antigen_demand.get(vac, [0] * FORECAST_PERIODS))
        for p_idx, val in enumerate(vals[:FORECAST_PERIODS]):
            apply_table_style(ws, row_num, p_idx + 2, val, is_alt=is_alt, formula=True)
        row_num += 1

    row_num += 1

    ws.cell(row=row_num, column=1, value='ANTIGEN-LEVEL AGGREGATION').font = \
        Font(bold=True, size=9, color=BLUE_FONT, name='Arial')
    ws.cell(row=row_num, column=1).fill = make_fill(SECTION_FILL)
    row_num += 1

    for antigen in ALL_ANTIGENS:
        is_alt = row_num % 2 == 0
        cell = ws.cell(row=row_num, column=1, value=antigen)
        cell.font = Font(bold=True, size=9, name='Arial')
        cell.fill = make_fill(SECTION_FILL if not is_alt else ALT_FILL)

        vals = antigen_demand.get(antigen, [0] * FORECAST_PERIODS)
        for p_idx, val in enumerate(vals[:FORECAST_PERIODS]):
            apply_table_style(ws, row_num, p_idx + 2, val, is_alt=is_alt, formula=True)
        row_num += 1

    ws.column_dimensions['A'].width = 30
    for i in range(2, FORECAST_PERIODS + 3):
        ws.column_dimensions[get_column_letter(i)].width = 16


def _build_medium_demand_sheet(wb, scenario_data, actual_md):
    ws = wb.create_sheet('Medium_Demand')
    ws.freeze_panes = 'B4'

    periods = list(range(1, FORECAST_PERIODS + 1))

    ws.merge_cells(_MERGE_RANGE)
    t = ws['A1']
    t.value = (f'Medium Demand Forecast – Antigen-Level '
               f'({FORECAST_PERIODS} Forecast Periods, '
               f'{FORECAST_START_YEAR}–{FORECAST_END_YEAR})')
    t.font = Font(bold=True, color='FFFFFFFF', size=12, name='Arial')
    t.fill = make_fill(HEADER_FILL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells(_MERGE_RANGE_ROW2)
    sub = ws['A2']
    sub.value = (f'Demand in doses. Columns 1–{FORECAST_PERIODS} represent forecast periods '
                 f'({FORECAST_START_YEAR}–{FORECAST_END_YEAR}). '
                 f'Antigens aggregated from contributing vaccine formats.')
    sub.font = Font(italic=True, size=9, color='FF595959', name='Arial')
    sub.alignment = Alignment(horizontal='left')

    ws.cell(row=3, column=1, value='Antigen').font = \
        Font(bold=True, color='FFFFFFFF', size=10, name='Arial')
    ws.cell(row=3, column=1).fill = make_fill(HEADER_FILL)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')
    for p in periods:
        cell = ws.cell(row=3, column=p + 1, value=f'Period {p}')
        cell.font = Font(bold=True, color='FFFFFFFF', size=10, name='Arial')
        cell.fill = make_fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal='center')

    source_antigens = list(actual_md.index) if actual_md is not None else []
    medium_data = scenario_data['Medium']

    row_num = 4
    display_antigens = source_antigens if source_antigens else list(medium_data.keys())

    for antigen in display_antigens:
        is_alt = row_num % 2 == 0
        cell = ws.cell(row=row_num, column=1, value=antigen)
        cell.font = Font(bold=True, size=10, name='Arial')
        cell.fill = make_fill(SECTION_FILL if not is_alt else ALT_FILL)
        cell.border = border_thin()

        if actual_md is not None and antigen in actual_md.index:
            vals = actual_md.loc[antigen].values[:FORECAST_PERIODS]
        else:
            vals = medium_data.get(antigen, [0] * FORECAST_PERIODS)

        for p_idx, val in enumerate(vals[:FORECAST_PERIODS]):
            apply_table_style(ws, row_num, p_idx + 2,
                              float(val) if pd.notna(val) else 0.0,
                              is_alt=is_alt, formula=True)
        row_num += 1

    # Totals row
    row_num += 1
    ws.cell(row=row_num, column=1, value='TOTAL (All Antigens)').font = \
        Font(bold=True, size=10, color=BLUE_FONT, name='Arial')
    ws.cell(row=row_num, column=1).fill = make_fill(SECTION_FILL)
    ws.cell(row=row_num, column=1).border = border_thin()

    start_data = 4
    end_data = row_num - 2
    for p in range(1, FORECAST_PERIODS + 1):
        col = p + 1
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}{start_data}:{col_letter}{end_data})'
        cell = ws.cell(row=row_num, column=col, value=formula)
        cell.font = Font(bold=True, size=10, color=BLACK_FONT, name='Arial')
        cell.fill = make_fill(SECTION_FILL)
        cell.number_format = '#,##0'
        cell.border = border_thin()
        cell.alignment = Alignment(horizontal='right')

    # Scenario summary block
    row_num += 3
    ws.cell(row=row_num, column=1, value='SCENARIO COMPARISON SUMMARY').font = \
        Font(bold=True, size=11, color=BLUE_FONT, name='Arial')
    ws.cell(row=row_num, column=1).fill = make_fill(SECTION_FILL)
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=FORECAST_PERIODS + 1)
    row_num += 1

    for p in periods:
        cell = ws.cell(row=row_num, column=p + 1, value=f'P{p}')
        cell.font = Font(bold=True, color='FFFFFFFF', size=9, name='Arial')
        cell.fill = make_fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=1, value='Scenario').font = \
        Font(bold=True, color='FFFFFFFF', size=9, name='Arial')
    ws.cell(row=row_num, column=1).fill = make_fill(HEADER_FILL)
    row_num += 1

    scenario_colors = {'Low': 'FFFFD700', 'Medium': 'FF90EE90', 'High': 'FFFFB6C1'}
    for scenario, factor in SCENARIOS.items():
        cell = ws.cell(row=row_num, column=1, value=f'{scenario} ({factor:.0%})')
        cell.font = Font(bold=True, size=9, name='Arial')
        cell.fill = make_fill(scenario_colors[scenario])
        cell.border = border_thin()

        for p_idx in range(FORECAST_PERIODS):
            total = sum(
                float(v) if pd.notna(v) else 0.0
                for antigen, vals in scenario_data[scenario].items()
                for i, v in enumerate([vals[p_idx]])
            )
            c = ws.cell(row=row_num, column=p_idx + 2, value=total)
            c.number_format = '#,##0'
            c.font = Font(size=9, name='Arial')
            c.fill = make_fill(scenario_colors[scenario])
            c.border = border_thin()
            c.alignment = Alignment(horizontal='right')
        row_num += 1

    ws.column_dimensions['A'].width = 22
    for i in range(2, FORECAST_PERIODS + 3):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Notes
    row_num += 2
    notes = [
        'Model Notes:',
        '• Demand = sum of projected supply capacity across all vaccine formats containing each antigen',
        '• Projections use weighted 5-year moving average with dampened linear trend extrapolation',
        '• Market shares applied per producer allocation table (sheet: Vaccine-Producer List)',
        '• Low scenario = Base × 0.85 | Medium = Base × 1.00 | High = Base × 1.15',
        f'• Source data: Demand_Scenarios_Base.xlsx | Periods 1–{FORECAST_PERIODS} = {FORECAST_START_YEAR}–{FORECAST_END_YEAR}',
    ]
    for note in notes:
        cell = ws.cell(row=row_num, column=1, value=note)
        cell.font = Font(italic=(not note.endswith(':')), size=8, name='Arial',
                         bold=note.endswith(':'))
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=FORECAST_PERIODS + 1)
        row_num += 1


# ─────────────────────────────────────────────────────────────────────────────
# 10. MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print('Loading source data...')
    sheets = load_source()

    print('Parsing capacity data...')
    cap_df = parse_capacity(sheets)
    print(f'  → {len(cap_df)} (vaccine, income group) series loaded')

    print('Building projections...')
    projections = build_capacity_projections(cap_df)

    print('Aggregating antigen demand...')
    antigen_demand = aggregate_demand_by_antigen(cap_df, projections)

    print('Applying scenarios...')
    scenario_data = apply_scenarios(antigen_demand)

    print('Loading actual Medium_Demand values...')
    actual_md = parse_medium_demand(sheets)
    print(f'  → {len(actual_md)} antigens loaded from source')

    print('Building Excel output...')
    wb = build_excel(sheets, cap_df, projections, antigen_demand, scenario_data, actual_md)

    out_path = 'Demand_Scenarios_Model.xlsx'
    wb.save(out_path)
    print(f'Saved: {out_path}')
    return out_path

if __name__ == '__main__':
    out = main()
    print(f'\nDone → {out}')
